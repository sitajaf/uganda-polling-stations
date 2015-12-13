from openpyxl import load_workbook


def convert():
    work_book = load_workbook('register.xlsx')
    work_book._active_sheet_index = 0

    work_sheet = work_book.active

    cell = work_sheet.cell('A4')
    cell2 = work_sheet.cell('E8')

    print('content in A4: ', cell.value)
    print('content in E8: ', cell2.value)


convert()
